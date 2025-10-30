from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import re
import pandas
import sys

# ! Константы таблицы, высчитанные в ручную !
# Константы недели
WEEK_LOOP_START = 4
WEEK_LOOP_LIMIT = 86
WEEK_LOOP_STEP = 14
# Константы дня
LECT_LOOP_START = 0
LECT_LOOP_END = 14

class schedule_reconciliation:
    # В конструкторе загружаем таблицы с нагрузкой и расписанием на основе полученных из командной строки путей. 
    def __init__(self, workload_path: str, schedule_path: str,week_loop_start: int, week_loop_limit: int, week_loop_step: int, lect_loop_start: int, lect_loop_end: int):
        self.workload = pandas.read_excel(workload_path, sheet_name="По ППС", usecols=["Мероприятие реестра, норма времени", "Вид потока", "План. поток", "ППС"], skiprows=1)
        self.schedule_workbook = load_workbook(schedule_path) # Загрузка таблицы с расписанием
        self.schedule = self.schedule_workbook[self.schedule_workbook.sheetnames[0]] # Загрузка первого листа таблицы
        self.mismatch_messages = []
        self.no_prof_messages = []
        self.week_loop_start = week_loop_start
        self.week_loop_limit = week_loop_limit
        self.week_loop_step = week_loop_step
        self.lect_loop_start = lect_loop_start
        self.lect_loop_end = lect_loop_end
        self.schedule_parser()

    # Метод для форматирования названия дисциплины
    def subject_formatting(self, subject: str):
        # Этап 1 - Избивляемся от переносов строки
        subject = subject.strip()
        subject = re.sub(r'\s+', ' ', subject)

        # Этап 2 - Вытаскиваем текст между номерами недель, и указанием подгруппы
        result = re.search(r"н.\s+([^\(]+)\s+\(",subject)
        if result is not None and len(result.group(1)) > 0:
            subject = result.group(1)

        # Этап 3 - Удаляем указание недель
        second_result = re.search(r"^([\d,]+ н\.)\s+(.*)$",subject)
        if second_result is not None and len(second_result.group(2)) > 0:
            subject = second_result.group(2)
        
        return subject
    # Метод для форматирования имени преподавателя
    def professor_formatting(self, professor_name: str,source: str):
        # Форматируем имя преподавателя, регулярное выражение меняется в зависимости от источника имени
        match source:
            # Форматировани имени из расписания
            case "schedule":
                result = re.search(r'^([А-Яа-яЁё]+)\s([А-Яа-яЁё])[а-яё]*\s([А-Яа-яЁё])\.?$',professor_name)
                if result is not None:
                    professor = result.group(1) + " " + result.group(2)+"."+result.group(3)+"."
                    return professor
                else:
                    return professor_name

            # Форматирование имени из нагрузки
            case "workload":
                result = re.search(r"^([А-Яа-яЁё]+)\s([А-Яа-яЁё])\w*\s([А-Яа-яЁё])\w",professor_name)
                if result is not None:
                    expected_professor = result.group(1) + " " + result.group(2)+"."+result.group(3)+"."
                    return expected_professor
                else:
                    pass
            case _:
                pass
    
    # Метод формирования сообщения об ошибки/несоответствии с нагрузкой
    def message_generator(self, group: str, subject: str, class_type: str, type:str, expected_professor = "", set_professor = ""):
        set_professor = set_professor or "Не указано"
        match type:
            case "no professor":
                no_prof_message = subject + " " + group + " " + class_type +  " НЕ УКАЗАН ПРЕПОДАВАТЕЛЬ"
                no_prof_message = no_prof_message.replace('\n', ' ').replace('\r', ' ')
                self.no_prof_messages.append(no_prof_message)
                pass
            case "mismatch":
                wrong_prof_message = subject + " " + group + " " + class_type + " ДОЛЖЕН БЫТЬ: " + expected_professor + " СТОИТ: " + set_professor
                self.mismatch_messages.append(wrong_prof_message)
                pass
            case _:
                pass

    # Метод сверки с нагрузкой
    def workload_matcher(self, group: str, subject:str, class_type:str, set_professor:str):
        # Фильтруем нагрузку до одной строки, с нужной группой, форматом занятия и дисциплиной
        subject_workload = self.workload[self.workload["Мероприятие реестра, норма времени"].str.contains(subject, na=False, case=False, regex=False) & self.workload["План. поток"].str.contains(group,na=False,case=False,regex=False) & self.workload["Вид потока"].str.contains(class_type, na=False)]
        # Проверка на присутствие в нагрузке проверяемой дисциплины
        if subject_workload.empty == True:
            pass
        # Если дисцпилина существует в нагрузке, продолжаем сверку
        else:
            expected_professor = self.professor_formatting(subject_workload["ППС"].iloc[0],"workload")
            # Если значение из нагрузки и расписания не сошлось, формируем сообщение о несоответствии и добавляем его в список
            if expected_professor != set_professor:
                self.message_generator(group,subject,class_type,"mismatch", expected_professor,set_professor)

    # Метод выявляющий столбцы в которых указаны группы и фиксирующий их название
    def group_finder(self):
        group_columns = []
        group_names = []
        # Метод для поиска столбцов с учебными группами. 
        for row in self.schedule.iter_rows():
            for cell in row:
                if cell.value != None:
                    if re.match(r'[A-Я]{4}-[0-9]{2}-[0-9]{2}',str(cell.value)):
                        group_columns.append(cell.column_letter)
                        group_names.append(cell.value)
        self.week_parser(group_columns,group_names)
    
    # Метод анализа расписания недели каждой учебной группы, найденной в group_finder
    def week_parser(self, group_columns: list, group_names: list):
        for group in group_columns:
            # Проход по недели (4,86,14)
            for week in range(self.week_loop_start,self.week_loop_limit,self.week_loop_step):
                # Проход по дням
                for lecture in range(self.lect_loop_start,self.lect_loop_end):
                    # Вытягиваем значение дисциплины из таблицы 
                    subject = self.schedule[group+str(week+lecture)].value 
                    # Если поле дисциплины пустое, то пропускаем строку
                    if subject is None: 
                        pass
                    else:
                        subject = self.subject_formatting(str(subject))
                        # Записываем данные о преподавателе и виде занятия
                        professor = self.schedule[get_column_letter(column_index_from_string(group)+2)+str(week+lecture)].value
                        class_type = self.schedule[get_column_letter(column_index_from_string(group)+1)+str(week+lecture)].value

                        # Если занятие присутствует, продолжаем проверку
                        if class_type == None:
                            pass

                        # Проверка на то указан ли преподаватель, если нет, то выдаем сообщение об отсутствии преподавателя
                        elif professor == None and class_type != None:
                            class_type = str(class_type)[:3] # Производим срез типа занятия для избежания многострочного вывода
                            self.message_generator(group_names[group_columns.index(group)], subject, class_type, type="no professor")
                        
                        # Если преподаватель указан, начинаем сверку с нагрузкой
                        else:
                            class_type = str(class_type)[:3] # Производим срез типа занятия для избежания многострочного вывода                        
                            self.workload_matcher(group_names[group_columns.index(group)],subject,class_type,self.professor_formatting(professor,"schedule"))
    
    # Функция запускающая прогон парсера и выводящая итоговое сообщение
    def schedule_parser(self):
        print("Добро пожаловать в программу сверки расписания")
        self.group_finder()
        if not self.mismatch_messages and not self.no_prof_messages:
            print("При сверке не обнаружено ошибок")
        elif not self.mismatch_messages and self.no_prof_messages:
            print("При сверке обнаружены следующие ошибки:")
            for message in self.no_prof_messages:
                print(message)
        elif self.mismatch_messages and not self.no_prof_messages:
            print("При сверке расписания обнаружены следующие расхождения:")
            for message in self.mismatch_messages:
                print(message)
        else:
            print("При сверке расписания обнаружены следующие ошибки:")
            for message in self.no_prof_messages:
                print(message)
            print()
            print("При сверке расписания обнаружены следующие расхождения:")
            for message in self.mismatch_messages:
                print(message)
                        
 
path_schedule = sys.argv[1]
path_workload = sys.argv[2]
schedule_parser = schedule_reconciliation(path_workload,path_schedule,WEEK_LOOP_START,WEEK_LOOP_LIMIT,WEEK_LOOP_STEP,LECT_LOOP_START,LECT_LOOP_END)
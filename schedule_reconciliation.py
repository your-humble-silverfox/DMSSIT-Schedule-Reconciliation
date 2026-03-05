from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import re
import pandas
import sys

# ! Manually calculated constants for excel spreadsheet calculation !
# Week loop constants
WEEK_LOOP_START = 4
WEEK_LOOP_LIMIT = 86
WEEK_LOOP_STEP = 14
# Day lecture loop constants 
LECT_LOOP_START = 0
LECT_LOOP_END = 14

class schedule_reconciliation:
    # Loading schedule and workload spreadsheets in constructor, as well as initializing variables for keeping generated messages on mismatches and errors
    def __init__(self, workload_file, schedule_file,week_loop_start=WEEK_LOOP_START, week_loop_limit=WEEK_LOOP_LIMIT, week_loop_step=WEEK_LOOP_STEP, lect_loop_start=LECT_LOOP_START, lect_loop_end=LECT_LOOP_END):
        #self.workload = pandas.read_excel(workload_path, sheet_name="По ППС", usecols=["Мероприятие реестра, норма времени", "Вид потока", "План. поток", "ППС"], skiprows=1)
        workload_file.seek(0)
        schedule_file.seek(0)
       
        self.workload = pandas.read_excel(workload_file, sheet_name="По ППС", usecols=["Мероприятие реестра, норма времени", "Вид потока", "План. поток", "ППС"], skiprows=1)
        #self.schedule_workbook = load_workbook(schedule_path) # Загрузка таблицы с расписанием
        self.schedule_workbook = load_workbook(schedule_file)
        self.schedule = self.schedule_workbook[self.schedule_workbook.sheetnames[0]] # Загрузка первого листа таблицы
        self.mismatch_messages = []
        self.no_prof_messages = []
        self.week_loop_start = week_loop_start
        self.week_loop_limit = week_loop_limit
        self.week_loop_step = week_loop_step
        self.lect_loop_start = lect_loop_start
        self.lect_loop_end = lect_loop_end

    # Method for formatting subject title
    def subject_formatting(self, subject: str):
        # First regex to remove any and all line breaks
        subject = subject.strip()
        subject = re.sub(r'\s+', ' ', subject)

        # Second regex to capture anything between week and sub-group information
        result = re.search(r"н.\s+([^\(]+)\s+\(",subject)
        if result is not None and len(result.group(1)) > 0:
            subject = result.group(1)

        # Third regex to remove week info
        second_result = re.search(r"^([\d,]+ н\.)\s+(.*)$",subject)
        if second_result is not None and len(second_result.group(2)) > 0:
            subject = second_result.group(2)
        
        return subject
    
    # Method for formatting professor names acquired from schedule file
    def schedule_professor_formatting(self, professor_name):
        result = re.search(r'^([А-Яа-яЁё]+)\s([А-Яа-яЁё])[а-яё]*\s([А-Яа-яЁё])\.?$',professor_name)
        if result is not None:
            professor = result.group(1) + " " + result.group(2)+"."+result.group(3)+"."
            return professor
        else:
            return professor_name
    
    # Method for formatting professor names acquired from workload file
    def workload_professor_formatting(self, professor_name):
        result = re.search(r"^([А-Яа-яЁё]+)\s([А-Яа-яЁё])\w*\s([А-Яа-яЁё])\w",professor_name)
        if result is not None:
            expected_professor = result.group(1) + " " + result.group(2)+"."+result.group(3)+"."
            return expected_professor
        else:
            pass
    
    # Method for generating a message in case of mismatch of professors in workload and schedule
    def mismatch_message_generator(self, group:str, subject: str, class_type:str, expected_professor:str, set_professor:str):
        set_professor = set_professor or "Не указано"
        message = (
            f"{subject} {group} {class_type} "
            f"ДОЛЖЕН БЫТЬ: {expected_professor} "
            f"СТОИТ: {set_professor}"
        )
        self.mismatch_messages.append(message)

    # Method for generating a message in case of no professor being set in the schedule
    def missing_professor_message_generator(self, group:str, subject: str, class_type:str):
        message = (
            f"{subject} {group} {class_type} "
            f"НЕ УКАЗАН ПРЕПОДАВАТЕЛЬ"
        )
        self.no_prof_messages.append(
            message.replace('\n', ' ').replace('\r', ' ')
        )

    # Method for matching data from schedule with workload
    def workload_matcher(self, group: str, subject:str, class_type:str, set_professor:str):
        # Filtering workload down to a single line with the necessary group, class type and subject
        subject_workload = self.workload[self.workload["Мероприятие реестра, норма времени"].str.contains(subject, na=False, case=False, regex=False) & self.workload["План. поток"].str.contains(group,na=False,case=False,regex=False) & self.workload["Вид потока"].str.contains(class_type, na=False)]
        # Checking if the subject is part of department's workload
        if subject_workload.empty:
            return
        # If the subject is a part of department's workload, continue the reconciliation
        else:
            #expected_professor = self.professor_formatting(subject_workload["ППС"].iloc[0],"workload")
            expected_professor = self.workload_professor_formatting(subject_workload["ППС"].iloc[0])
            # If the professor attached to the lesson in workload didn't match the one stated in schedule, form a message and add it to the message list
            if expected_professor != set_professor:
                self.mismatch_message_generator(group,subject,class_type, expected_professor, set_professor)

    # Method for finding columns, containing group schedules
    def group_finder(self):
        group_columns = []
        group_names = []
        for row in self.schedule.iter_rows():
            for cell in row:
                if cell.value != None:
                    if re.match(r'[A-Я]{4}-[0-9]{2}-[0-9]{2}',str(cell.value)):
                        group_columns.append(cell.column_letter)
                        group_names.append(cell.value)
        self.week_parser(group_columns,group_names)
    
    # Method for analyzing each groups week schedule
    def week_parser(self, group_columns: list, group_names: list):
        for group in group_columns:
            # Loop going through weeks
            for week in range(self.week_loop_start,self.week_loop_limit,self.week_loop_step):
                # Loop going through individual subjects
                for lecture in range(self.lect_loop_start,self.lect_loop_end):
                    # Pulling subject name from the schedule spreadsheet
                    subject = self.schedule[group+str(week+lecture)].value 
                    # If there is none, skipping onto the next
                    if subject is None: 
                        pass
                    else:
                        # Formatting the subject, as well as saving professor and class-type (lecture, seminary, lab work) for further analysis
                        subject = self.subject_formatting(str(subject))
                        professor = self.schedule[get_column_letter(column_index_from_string(group)+2)+str(week+lecture)].value
                        class_type = self.schedule[get_column_letter(column_index_from_string(group)+1)+str(week+lecture)].value

                        # Due to the specifics of the schedule table, we have to ensure that subject is not present by also verifying that it's class_type is absent in the spreadsheet
                        if class_type == None:
                            pass

                        # Checking for no professor listed at the lesosn
                        elif professor == None and class_type != None:
                            class_type = str(class_type)[:3] # Slicing class-type to ensure a one-line message
                            self.missing_professor_message_generator(group_names[group_columns.index(group)], subject, class_type)
                        
                        # If the professor is set, comparing the entry with the workload
                        else:
                            class_type = str(class_type)[:3] # Slicing class-type to ensure a one-line message                       
                            self.workload_matcher(group_names[group_columns.index(group)],subject,class_type,self.schedule_professor_formatting(professor))
    
    # Method for full schedule reconciliation
    def full_check(self):
        self.group_finder()
        if not self.mismatch_messages and not self.no_prof_messages:
            return {
                "status": "ok"
            }
        else:
            return {
                "status": "errors_found",
                "missing_professors": self.no_prof_messages,
                "mismatched_professors": self.mismatch_messages
            }
    
    # Method for reconciliation, returning only mismatches
    def mismatch_check(self):
        self.group_finder()

        if not self.mismatch_messages:
            return{
                "status":("No mismatch found")
            }
        else:
            return{
                "status":("Mismatches_found"),
                "mismatched_professors": self.mismatch_messages
            }
    
    # Method for reconciliation, returning specifically missing professors
    def no_prof_check(self):
        self.group_finder()

        if not self.no_prof_messages:
            return{
                "status":("Missing professors not found")
            }
        else:
            return{
                "status":("Missing Professors found"),
                "mismatched_professors": self.no_prof_messages
            }
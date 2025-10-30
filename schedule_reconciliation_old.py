from openpyxl import load_workbook
import re
import pandas
import sys

def startup():
    print("Добро пожаловать в программу сверки расписания")
    path_schedule = input("Введите путь к таблице с расписание в формате .xlxs: ")
    path_workload = input("Введите путь к таблице с нагрузкой в формате .xls: ")
    schedule_parser(path_schedule, workload_parser(path_workload))

def workload_parser(path: str):
    workload = pandas.read_excel(path, sheet_name="По ППС", usecols=["Мероприятие реестра, норма времени", "Вид потока", "План. поток", "ППС"], skiprows=1)
    workload["Мероприятие реестра, норма времени"] = workload["Мероприятие реестра, норма времени"].astype(str).str.strip()
    workload["Вид потока"] = workload["Вид потока"].str.strip()
    return workload

def num_to_column(num: int):
    column = ""
    while (num > 0):
        num -= 1
        column = chr( num % 26 + ord('A')) + column
        num //= 26
    return column

def column_to_num(column: str):
    # ! USELESS CODE !
    num = 0
    for c in column:
        num = num * 26 + (ord(c.upper()) - ord('A') + 1)
    return num

# ? Should it be rewritten with pandas ?
def schedule_parser(path: str, workload: pandas.DataFrame):
    # TODO: Переписать с хардкода значений на поиск значений в таблице
    # ! Deprecated Code !
    # Загружаем таблицу и выбираем первый лист
    wb = load_workbook(path)
    sheet = wb[wb.sheetnames[0]]
    # При помощи вложенных циклов и условий определяем столбцы групп и начинаем их парсить
    # Шагаем по ячейкам где располагаются названия групп
    for i in range(6, 10000 ,5):
        if sheet[num_to_column(i)+"2"].value == None:
            break
        elif sheet[num_to_column(i)+"2"].value != "День недели":
            group = sheet[num_to_column(i)+"2"].value
            # Проход по недели
            for j in range(4,86,14):
                # Проход по дням
                for k in range(0,14):
                    subject = sheet[num_to_column(i)+str(k+j)].value # Вытягиваем значение предмета из таблицы 
                    if subject is None: # Если поле дисциплины пустое, то пропускаем строку
                        pass
                    else:
                        # Регуляркой форматируем дисциплину до однострочного наименования дисциплины
                        subject = str(subject)
                        subject = re.sub(r'\s+', ' ', subject.strip())
                        result = re.search(r"н.\s+([^\(]+)\s+\(",subject)
                        if result is not None and len(result.group(1)) > 0:
                            subject = result.group(1)
                        second_result = re.search(r"^([\d,]+ н\.)\s+(.*)$",subject)
                        if second_result is not None and len(second_result.group(2)) > 0:
                            subject = second_result.group(2)
                        # Записываем данные о преподавателе и виде занятия
                        professor = sheet[num_to_column(i+2)+str(k+j)].value
                        class_type = sheet[num_to_column(i+1)+str(k+j)].value
                        class_type = str(class_type)[:3] # Производим срез типа занятия для избежания многострочного вывода
                        # Если занятие присутствует, продолжаем проверку
                        if class_type == "Non":
                            pass
                        # Проверка на то указан ли преподаватель, если нет, то выдаем сообщение об отсутствии преподавателя
                        elif professor == None and class_type != None:
                            no_prof_message = subject + " " + group + " " + class_type +  " НЕ УКАЗАН ПРЕПОДАВАТЕЛЬ"
                            no_prof_message = no_prof_message.replace('\n', ' ').replace('\r', ' ')
                            print(no_prof_message)
                        # Если преподаватель указан, начинаем сверку с нагрузкой
                        else:
                            # При помощи регулярного выражения приводим имя преподвателя из расписания к формату Фамилия И.О.
                            result = re.search(r'^([А-Яа-яЁё]+)\s([А-Яа-яЁё])[а-яё]*\s([А-Яа-яЁё])\.?$',professor)
                            if result is not None:
                                professor = result.group(1) + " " + result.group(2)+"."+result.group(3)+"."
                            # Фильтруем нагрузку до одной строки, с нужной группой, форматом занятия и дисциплиной
                            subject_workload = workload[workload["Мероприятие реестра, норма времени"].str.contains(subject, na=False, case=False, regex=False) & workload["План. поток"].str.contains(group,na=False,case=False,regex=False) & workload["Вид потока"].str.contains(class_type, na=False)]
                            # Проверка на то, есть ли дисциплина в нагрузке вообще
                            if subject_workload.empty == True:
                                pass
                            # Если дисцпилина существует в нагрузке, продолжаем сверку
                            else:
                                # Достаем из нагрузки ожидаемое имя преподавателя и форматируем его до Фамилия И.О.
                                expected_professor = subject_workload["ППС"].iloc[0]
                                result = re.search(r"^([А-Яа-яЁё]+)\s([А-Яа-яЁё])\w*\s([А-Яа-яЁё])\w",expected_professor)
                                if result is not None:
                                    expected_professor = result.group(1) + " " + result.group(2)+"."+result.group(3)+"."
                                    # Если значение из нагрузки и расписания не сошлось, формируем и выводим сообщение о несоответствии
                                    if expected_professor != professor:
                                        wrong_prof_message = subject + " " + group + " " + class_type + " ДОЛЖЕН БЫТЬ: " + expected_professor + " СТОИТ: " + professor
                                        print(wrong_prof_message)
                        
        else:
            pass

print("Добро пожаловать в программу сверки расписания")
path_schedule = sys.argv[1]
path_workload = sys.argv[2]
schedule_parser(path_schedule, workload_parser(path_workload))

#workload_parser("Учебная_нагрузка_читающего_подразделения_форма_A_02_09_2025.xls")
# startup()

#schedule_parser("ИИТ_2 курс_25-26_осень.xlsx",workload_parser("test_workload.xls"))